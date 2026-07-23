"""
Vincore AI — TEMPLATE: Gantt Chart / Project Schedule Build Script
──────────────────────────────────────────────────────────────────
HOW TO USE:
  1. Edit client_config.json for the new client
  2. Edit the MILESTONES and TASKS lists below to match the new project
  3. Run:  python3 build_gantt_template.py
  4. Output file appears in ./output/

This script is the reference template for all Python/Excel build scripts.
The pattern is identical for build_raid.py and build_gantt.py:
  - C['CLIENT'], C['CODE'], C['NAME'] etc. replace every hardcoded MBH value
  - C['filename']('gantt', 'xlsx') builds the output filename
"""

import os
import sys

# ── Import config (the only import change needed) ──────────────────
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config_loader import C

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── COLORS (unchanged across all clients) ─────────────────────────
NAVY='081C3A'; BLUE='2196F3'; CYAN='54E0FF'; WHITE='FFFFFF'
BODY='37415f'; MUTED='6b7794'; PAPER='F2F9FF'; SILVER='C8D3E6'
GREEN='10B981'; AMBER='D97706'; RED='EF4444'

def F(h): return PatternFill('solid', fgColor=h)
def fnt(sz=10, bold=False, color=BODY):
    return Font(name='Arial', size=sz, bold=bold, color=color)
def bdr():
    s = Side(style='thin', color=SILVER)
    return Border(top=s, bottom=s, left=s, right=s)
def aln(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ── PROJECT-SPECIFIC DATA ─────────────────────────────────────────
# EDIT THESE for each new client project. Keep the same column structure.
# (phase, wbs_id, task_name, type, owner, start_day, end_day, effort)

TASKS = [
    # ── CUSTOMISE: Replace MBH tasks with new client's WBS ────────
    # Examples — edit to match the new project:
    ('Phase 1 — Initiate',  None, f'PHASE 1 — INITIATE', 'PHASE_HDR', None, None, None, None),
    ('Phase 1 — Initiate',  None, '★  Milestone 1: Commencement', 'MILESTONE', 'Both', 0, 0, None),
    ('Phase 1 — Initiate', '1.1', 'Client Discovery & Documentation', 'TASK', C['vincore']['founder'], 0, 1, '2 hrs'),
    ('Phase 1 — Initiate', '1.2', 'Commercial Documents', 'TASK', C['vincore']['founder'], 0, 1, '2 hrs'),
    ('Phase 1 — Initiate', '1.3', 'Agreement Execution & Deposit Collection', 'TASK', 'Both', 0, 0, '0.5 hrs'),
    ('Phase 1 — Initiate', '1.4', 'Kickoff Preparation', 'TASK', C['vincore']['founder'], 0, 1, '2 hrs'),
    ('Phase 1 — Initiate', '1.5', 'Kickoff Meeting with Client', 'TASK', 'Both', 1, 2, '1.5 hrs'),
    ('Phase 1 — Initiate',  None, '★  Milestone 2: Kickoff Complete', 'MILESTONE', 'Both', 2, 2, None),

    # ADD REMAINING PHASES BELOW following the same pattern
    # Copy from build_gantt.py (MBH version) and adjust task names,
    # days, and efforts for the new client
]

MILESTONES = [
    # (id, name, phase, day, owner, success_condition)
    ('M1', 'Commencement',    'Initiate', 'Day 0',    'Both',          f'Signed agreement and {C["DEPOSIT"]} deposit received'),
    ('M2', 'Kickoff Complete','Initiate', 'Day 1–2',  'Both',          'Kickoff meeting held; brand assets and access received'),
    # ADD MORE MILESTONES below matching the new project
]

# ── BUILD WORKBOOK ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Project Schedule'
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'H6'

# ── BANNER ────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 38
ws.merge_cells('A1:V1')
c = ws['A1']
# Uses C['CLIENT'] and C['CODE'] from config — no hardcoding
c.value = f'VINCORE AI  |  PROJECT SCHEDULE & GANTT CHART  |  {C["CODE"]}  |  {C["CLIENT"]}'
c.font  = Font(name='Arial', size=14, bold=True, color=CYAN)
c.fill  = F(NAVY); c.alignment = aln('left','center')

ws.row_dimensions[2].height = 18
ws.merge_cells('A2:V2')
ws['A2'].value = (
    f'Document ID: {C["docs"]["gantt"]}  |  Client: {C["CLIENT"]}  |  '
    f'Project: {C["NAME"]}  |  Version: {C["VERSION"]}  |  {C["MONTH"]}'
)
ws['A2'].font  = Font(name='Arial', size=9, color=MUTED)
ws['A2'].fill  = F('0d1f38'); ws['A2'].alignment = aln('left','center')

# ── (Continue building Gantt table using TASKS list above) ─────────
# Copy the table-building logic from build_gantt.py (the MBH version)
# and reference TASKS defined above. No client values need to change —
# they are all pulled from C['CLIENT'], C['CODE'] etc. via the config.

# ── SAVE ──────────────────────────────────────────────────────────
os.makedirs(C['OUT'], exist_ok=True)
out_path = os.path.join(C['OUT'], C['filename']('gantt', 'xlsx', 'ProjectSchedule'))
wb.save(out_path)
print(f'✓  Built: {out_path}')
